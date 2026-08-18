# views_backup.py
#
# Exportar / importar TODA la DB del área de software (VPS, catálogo de
# servicios, instalaciones y scripts) como medida de seguridad — un
# backup/restore point antes de hacer cambios grandes o pruebas en el
# entorno real.
#
# Formato: Excel (.xlsx), una hoja por tabla, para poder abrirlo y
# revisarlo directamente. Por dentro se apoya en los serializers de
# Django (formato "python") para no reinventar la conversión de tipos
# ni perder la fecha de alta/modificación original al reimportar
# (Django preserva esos campos en modo "raw" al deserializar).
#
# No incluye los archivos subidos en Script.archivo (solo la referencia
# a la ruta): el backup es de los datos de la DB, no del filesystem.
#
# Restringido a superusuarios: reemplaza TODOS los datos del área al
# importar, es una operación destructiva por diseño (restaura un punto
# anterior), no un permiso de uso diario.

import datetime

from django.core import serializers
from django.db import models as djmodels, transaction
from django.http import HttpResponse, JsonResponse
from django.utils import timezone
from django.views import View
from openpyxl import Workbook, load_workbook

from .models import VPS, ServicioSoftware, Instalacion, Script

# Orden con dependencias resueltas: VPS/Servicios primero (sin FKs entre
# sí), Instalaciones/Scripts después (referencian a las anteriores).
HOJAS = {
    'VPS': VPS,
    'Servicios': ServicioSoftware,
    'Instalaciones': Instalacion,
    'Scripts': Script,
}


def _nombres_campos(modelo):
    return [f.name for f in modelo._meta.fields if not f.primary_key]


def _celda_para_exportar(valor):
    # Excel no admite datetimes "aware": los pasamos a hora local sin tz.
    if isinstance(valor, datetime.datetime) and timezone.is_aware(valor):
        return timezone.localtime(valor).replace(tzinfo=None)
    return valor


def _construir_workbook():
    wb = Workbook()
    wb.remove(wb.active)
    for hoja, modelo in HOJAS.items():
        ws = wb.create_sheet(hoja)
        campos = _nombres_campos(modelo)
        ws.append(['id'] + campos)
        registros = serializers.serialize('python', modelo.objects.all())
        for registro in registros:
            fila = [registro['pk']]
            for campo in campos:
                fila.append(_celda_para_exportar(registro['fields'][campo]))
            ws.append(fila)
    return wb


class SoftwareExportarView(View):
    def get(self, request):
        if not request.user.is_authenticated or not request.user.is_superuser:
            return JsonResponse({'error': 'Sin permiso'}, status=403)

        wb = _construir_workbook()
        filename = f"backup_software_{timezone.now():%Y%m%d_%H%M%S}.xlsx"
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        wb.save(response)
        return response


def _valor_para_importar(campo, valor):
    if valor is None:
        return None if campo.null else ''

    if isinstance(campo, (djmodels.ForeignKey, djmodels.IntegerField)) and isinstance(valor, float) and valor.is_integer():
        valor = int(valor)

    if isinstance(campo, djmodels.DateTimeField) and isinstance(valor, datetime.datetime) and timezone.is_naive(valor):
        return timezone.make_aware(valor, timezone.get_current_timezone())

    return valor


def _leer_workbook(wb):
    """Convierte el workbook a la estructura que espera serializers.deserialize('python', ...)."""
    datos_por_modelo = {}
    for hoja, modelo in HOJAS.items():
        ws = wb[hoja]
        filas = list(ws.iter_rows(values_only=True))
        registros = []
        if filas:
            campos = [str(c) for c in filas[0][1:]]
            for fila in filas[1:]:
                if fila[0] is None:
                    continue
                fields = {}
                for nombre_campo, valor in zip(campos, fila[1:]):
                    campo = modelo._meta.get_field(nombre_campo)
                    fields[nombre_campo] = _valor_para_importar(campo, valor)
                registros.append({
                    'model': f'{modelo._meta.app_label}.{modelo._meta.model_name}',
                    'pk': int(fila[0]),
                    'fields': fields,
                })
        datos_por_modelo[modelo] = registros
    return datos_por_modelo


class SoftwareImportarAjax(View):
    def post(self, request):
        if not request.user.is_authenticated or not request.user.is_superuser:
            return JsonResponse({'error': 'Sin permiso'}, status=403)

        archivo = request.FILES.get('archivo')
        if not archivo:
            return JsonResponse({'success': False, 'error': 'Subí un archivo de backup (.xlsx).'}, status=400)

        try:
            wb = load_workbook(archivo, data_only=True)
        except Exception:
            return JsonResponse({'success': False, 'error': 'El archivo no es un backup válido (.xlsx).'}, status=400)

        if set(HOJAS) - set(wb.sheetnames):
            return JsonResponse({
                'success': False,
                'error': 'El archivo no tiene el formato esperado del backup del área de software.',
            }, status=400)

        try:
            datos_por_modelo = _leer_workbook(wb)
        except Exception:
            return JsonResponse({'success': False, 'error': 'No se pudo leer el contenido del backup.'}, status=400)

        total = sum(len(registros) for registros in datos_por_modelo.values())
        if total == 0:
            return JsonResponse({'success': False, 'error': 'El backup no tiene registros.'}, status=400)

        try:
            objetos_por_modelo = {
                modelo: list(serializers.deserialize('python', registros))
                for modelo, registros in datos_por_modelo.items()
            }
        except Exception:
            return JsonResponse({'success': False, 'error': 'El backup tiene datos inválidos o corruptos.'}, status=400)

        with transaction.atomic():
            # Orden seguro de borrado: Instalacion/Script primero (dependen
            # de VPS/ServicioSoftware vía FK, Instalacion los protege).
            Instalacion.objects.all().delete()
            Script.objects.all().delete()
            VPS.objects.all().delete()
            ServicioSoftware.objects.all().delete()

            for modelo in HOJAS.values():
                for obj in objetos_por_modelo[modelo]:
                    obj.save()

        return JsonResponse({'success': True, 'cantidad': total})
